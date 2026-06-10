"""Rendezett MHS klub Excel sablon generálása a tisztított adatokkal."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

wb = Workbook()
ws = wb.active
ws.title = "Klubok"

headers = ["Klub", "Helyszín neve", "Település", "Cím", "Nap", "Időpont",
           "Link", "Logó URL", "Lat", "Lng"]

# Tisztított adatok: egy sor = egy edzéshelyszín (egy pötty)
rows = [
    ["Just Another HEMA Club", "Nagymező utcai terem", "Budapest",
     "Budapest, Nagymező u. 68, 1065", "Kedd", "18:00–20:00",
     "https://www.facebook.com/share/199MFcStLg/", "", "", ""],
    ["Just Another HEMA Club", "Huba utcai terem", "Budapest",
     "Budapest, Huba u. 7, 1134", "Péntek", "19:30–21:00",
     "https://www.instagram.com/just.another.hema.club", "", "", ""],
    ["Pécsi Hosszúkardvívó Sportkör", "", "Pécs",
     "Pécs, Szigeti út 12, 7624", "Hétfő, Csütörtök", "16:30–18:00 / 19:30–21:00",
     "https://www.facebook.com/profile.php?id=100057515436147", "", "", ""],
    ["Felvidéki Szablyavívó Iskola", "", "Komárno (SK)",
     "Eötvösova 3195/21, 945 05 Komárno, Szlovákia", "Hétfő, Szerda", "19:00–20:30",
     "https://www.facebook.com/felfoldiszablyavivoszovetseg", "", "", ""],
    ["Falka Vívóiskola", "", "Komárno (SK)",
     "Ulica Práce 461, 945 01 Komárno, Szlovákia", "Péntek", "17:30–19:30",
     "https://www.falka.sk", "", "", ""],
    ["Blood and Iron Martial Arts Hungary", "", "Szombathely",
     "Szombathely, Kötő u. 27, 9700", "Kedd, Péntek", "19:00–20:30",
     "https://www.facebook.com/profile.php?id=61565459563283", "", "", ""],
    ["Anjou Udvari Lovagkör Kardvívó és Önvédelmi Iskola SE", "", "Debrecen",
     "Debrecen, Simonffy u. 21, 4025", "Szombat", "14:00–16:00",
     "http://www.aule.hu/", "", "", ""],
    ["Kard Rendje VISE Budapest", "Csik Ferenc Gimnázium", "Budapest",
     "Budapest, Medve utca 5-7, 1027", "Hétfő, Szerda", "19:00–21:00 / 18:00–20:00",
     "https://www.krvise.com/budapest", "", "", ""],
    ["Kard Rendje VISE Győr", "Pattantyús-Ábrahám Géza Technikum", "Győr",
     "Győr, Ikva u. 70, 9024", "Hétfő, Csütörtök", "18:00–20:00",
     "https://www.krvise.com/gyor", "", "", ""],
    ["Ars Ensis Lovagi Kör és Kardvívó Iskola Egyesület", "Városligeti Sportcentrum", "Budapest",
     "Budapest, Olof Palme sétány 5, 1146", "Hétfő", "19:00–21:00",
     "https://www.arsensis.hu/edzesek", "", "", ""],
    ["Ars Ensis Lovagi Kör és Kardvívó Iskola Egyesület", "Kölcsey Ferenc Általános Iskola", "Győr",
     "Győr, Szent Imre út 33, 9024", "Csütörtök", "18:00–19:30",
     "https://www.arsensis.hu/edzesek", "", "", ""],
    ["Ars Ensis Lovagi Kör és Kardvívó Iskola Egyesület", "Noszlopy Gáspár Gimnázium és Kollégium", "Veszprém",
     "Veszprém, Tüzér u. 42, 8200", "Kedd", "19:00–20:30",
     "https://www.arsensis.hu/edzesek", "", "", ""],
]

# Fejléc stílus
header_fill = PatternFill("solid", start_color="3A2A1A")
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="C8B89A")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.append(headers)
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

# Megjegyzések a kulcsoszlopokhoz
ws.cell(row=1, column=9).comment = Comment(
    "Hagyd ÜRESEN, ha a szkript automatikusan keresse meg a koordinátát a címből. "
    "Csak akkor írj ide számot, ha pontosítani akarsz (pl. 47.4979). Akkor ezt használja.", "MHS")
ws.cell(row=1, column=10).comment = Comment(
    "Hagyd ÜRESEN auto-kereséshez. Kézi felülíráshoz írd be a hosszúsági fokot (pl. 19.0402).", "MHS")
ws.cell(row=1, column=2).comment = Comment(
    "Opcionális. Akkor hasznos, ha egy klubnak több terme van — így megkülönböztethető a térképen.", "MHS")

# Adatsorok
data_font = Font(name="Arial", size=10)
for r in rows:
    ws.append(r)
for row in ws.iter_rows(min_row=2, max_row=len(rows)+1, max_col=len(headers)):
    for c in row:
        c.font = data_font
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = border

# Oszlopszélességek
widths = [34, 28, 14, 40, 18, 24, 40, 24, 10, 10]
from openpyxl.utils import get_column_letter
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

wb.save("/home/claude/hema-terkep/MHS_klubok.xlsx")
print(f"Sablon kész: {len(rows)} helyszín, {len(headers)} oszlop")
