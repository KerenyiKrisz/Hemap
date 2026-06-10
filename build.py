#!/usr/bin/env python3
"""
MHS Interaktív Térkép – adat build szkript
==========================================

Beolvassa a MHS_klubok.xlsx fájlt, és legenerálja a data/clubs.json-t,
amiből a térkép dolgozik.

Hibrid geokódolás:
  - Ha a Lat/Lng oszlop ki van töltve az Excelben  -> azt használja (kézi felülírás).
  - Ha üresen van hagyva                            -> automatikusan lekéri a címből
                                                       (OpenStreetMap Nominatim, ingyenes).

Az egyszer már megtalált koordinátákat elmenti egy gyorsítótárba
(scripts/geocache.json), így legközelebb nem kell újra lekérni őket,
és nem terheled feleslegesen a szolgáltatást.

Használat:
    python scripts/build.py

Függőség:
    pip install openpyxl requests
"""

import json
import time
import sys
from pathlib import Path

try:
    import requests
    from openpyxl import load_workbook
except ImportError:
    print("Hiányzó csomag. Telepítsd:  pip install openpyxl requests")
    sys.exit(1)

# --- Útvonalak (a szkript helyétől függetlenül működik) ---
ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "MHS_klubok.xlsx"
OUT = ROOT / "data" / "clubs.json"
CACHE = ROOT / "scripts" / "geocache.json"

NOMINATIM = "https://nominatim.openstreetmap.org/search"
HEADERS = {"User-Agent": "MHS-HEMA-Terkep/1.0 (hosszukardvivas.hu)"}

# Oszlopnevek az Excelben -> JSON kulcsok
COLS = {
    "Klub": "club",
    "Helyszín neve": "venue",
    "Település": "city",
    "Cím": "address",
    "Nap": "days",
    "Időpont": "time",
    "Link": "link",
    "Logó URL": "logo",
    "Lat": "lat",
    "Lng": "lng",
}


def load_cache():
    if CACHE.exists():
        return json.loads(CACHE.read_text(encoding="utf-8"))
    return {}


def save_cache(cache):
    CACHE.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def geocode(address, cache):
    """Cím -> (lat, lng). Cache-eli az eredményt. None, ha nem található."""
    if address in cache:
        return cache[address]
    try:
        r = requests.get(
            NOMINATIM,
            params={"q": address, "format": "json", "limit": 1},
            headers=HEADERS,
            timeout=15,
        )
        r.raise_for_status()
        data = r.json()
        if data:
            lat = round(float(data[0]["lat"]), 6)
            lng = round(float(data[0]["lon"]), 6)
            cache[address] = [lat, lng]
            save_cache(cache)
            time.sleep(1.1)  # Nominatim udvariassági szabály: max 1 kérés/mp
            return [lat, lng]
    except Exception as e:
        print(f"   ! Geokódolási hiba: {e}")
    return None


def main():
    if not XLSX.exists():
        print(f"Nem találom: {XLSX}")
        sys.exit(1)

    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Klubok"]

    header = [c.value for c in ws[1]]
    idx = {name: header.index(name) for name in COLS if name in header}

    cache = load_cache()
    clubs = []
    geocoded, manual, failed = 0, 0, 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[idx["Klub"]]:
            continue

        rec = {}
        for hu, key in COLS.items():
            if hu in idx:
                val = row[idx[hu]]
                rec[key] = str(val).strip() if val is not None else ""

        # Koordináták: kézi felülírás vagy auto geokódolás
        lat = rec.get("lat", "")
        lng = rec.get("lng", "")
        if lat and lng:
            rec["lat"] = float(lat)
            rec["lng"] = float(lng)
            manual += 1
        else:
            print(f" -> Geokódolás: {rec['club']} ({rec['address']})")
            coords = geocode(rec["address"], cache)
            if coords:
                rec["lat"], rec["lng"] = coords[0], coords[1]
                geocoded += 1
            else:
                print(f"   !! Nem található koordináta, kihagyva: {rec['club']}")
                failed += 1
                continue

        clubs.append(rec)

    OUT.parent.mkdir(exist_ok=True)
    OUT.write_text(
        json.dumps(clubs, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print("\n" + "=" * 50)
    print(f"  Kész: {len(clubs)} helyszín a clubs.json-ban")
    print(f"  Kézi koordináta: {manual} | Auto-geokódolt: {geocoded} | Sikertelen: {failed}")
    print(f"  Kimenet: {OUT}")
    print("=" * 50)


if __name__ == "__main__":
    main()
